"""
Importer for 'Haushalt 2026.xlsx' -> local Postgres (doewe_local).

Modes:
  --dry-run   : print a full plan, write nothing
  --apply     : execute the plan inside a single DB transaction

Scope: Account acc_demo (user fam.thiemann@gmail.com), months Jan-Apr 2026.
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime

import psycopg2
import psycopg2.extras
from openpyxl import load_workbook

XLSX_PATH = "/Users/konrad.thiemann/Downloads/Haushalt 2026.xlsx"
DB_URL = "postgresql://doewe:doewe@localhost:5432/doewe_local"
ACCOUNT_ID = "acc_demo"
USER_EMAIL = "fam.thiemann@gmail.com"
YEAR = 2026
MONTH_COLS = {1: "C", 2: "D", 3: "E", 4: "F"}  # Jan–Apr only
DEFAULT_DAY = 15

# Row → (ExcelLabel, DbCategory, isIncome, descPrefix|None)
INCOME_ROWS = {
    2: ("Gehalt K", "Salary 1", True, None),
    3: ("Gehalt C", "Salary 2", True, None),
    4: ("Kindergeld", "Child benefit", True, None),
    5: ("Sonstiges (Einn.)", "Misc Income", True, None),
}

# Alltagsausgaben (Z. 8–26)
OUTCOME_ROWS = {
    8: ("Lebensmittel", "Groceries", False, None),
    9: ("Drogerie/Haushalt", "Drugstore", False, None),
    10: ("Kosmetik", "Cosmetics", False, None),
    11: ("Bestellen", "Food order", False, None),
    12: ("Essen gehen", "Eating out", False, None),
    13: ("Ausgehen", "Going out", False, None),
    14: ("Kleidung Charlie", "Clothing", False, "Charlie"),
    15: ("Kleidung Konni", "Clothing", False, "Konni"),
    16: ("Einkäufe Milan", "Kids – Milan", False, None),
    17: ("Einkäufe Liana", "Kids – Liana", False, None),
    18: ("Freizeit", "Leisure", False, None),
    19: ("Hobbies", "Hobbies", False, None),
    20: ("Interior", "Interior", False, None),
    21: ("Geschenke", "Presents", False, None),
    22: ("Essen unterwegs", "Food on-the-go", False, None),
    23: ("Gesundheit", "Health", False, None),
    24: ("Mobilität/Tanken", "Mobility", False, None),
    25: ("Besonderes", "Special", False, None),
    26: ("Sonstiges", "Misc", False, None),
}

# Jährliche Posten (Z. 29–36) → Fixed costs, only rows with >0 values.
YEARLY_ROWS = {
    29: "Golf Steuer",
    30: "GEZ",
    31: "Prime",
    32: "ADAC",
    33: "Mieterschutz",
    34: "Schule Förderverein",
    35: "Mietkaution",
    36: "BZ Abo",
}

# Sparen (Z. 38–41) → Savings, all with prefix for filtering
SAVINGS_ROWS = {
    38: ("Sparen", "Sparen"),
    39: ("Mili", "Mili"),
    40: ("Liana", "Liana"),
    41: ("Charlie Depot", "Charlie Depot"),
}

# Monatliche Fixkosten (Z. 43–66). Value "recurring" means the latest (April=F) value
# becomes a RecurringTransaction; any earlier month that deviates is booked as a single TX.
# "single" means KEIN Recurring, every month individually.
# "skip_zero" means if value == 0, no TX.
FIXED_COST_ROWS = {
    43: ("Miete", "recurring"),
    44: ("Strom & Gas", "recurring"),
    45: ("Internet / TV", "recurring"),
    46: ("Handy K", "recurring"),
    47: ("Handy C", "recurring"),
    48: ("audible", "skip_zero"),
    49: ("Kindle unlimited", "recurring"),
    50: ("Spotify family", "recurring"),
    51: ("Hausrat", "recurring"),
    52: ("Zahnzusatz K", "recurring"),
    53: ("Zahnzusatz C", "recurring"),
    54: ("Haftpflicht", "recurring"),
    55: ("BU C", "recurring"),
    56: ("Rechtsschutz", "recurring"),
    57: ("KFZ-Vers.", "single"),          # user: no recurring
    58: ("apollo", "recurring"),
    59: ("Verdi", "recurring"),
    60: ("Hort", "recurring"),
    61: ("Kita", "recurring"),
    62: ("Essensgeld M.", "recurring"),
    63: ("I-cloud", "recurring"),
    64: ("Internetseit Konni", "recurring"),
    65: ("Hansefit", "recurring"),
    66: ("Kreditkarte->tolino", "recurring"),
}

FIXED_COST_CATEGORY = "Fixed costs"


# ---------- Parsing ----------

# "16€ Edeka", "77,5€ Flink", "1255,60€ Krankengeld", "64€", "Spiegelburg, 7€ Fred"
AMOUNT_RE = re.compile(r"(?P<amount>\d{1,6}(?:[.,]\d{1,2})?)\s*€")
# Typo fallback: line begins with number followed by ) & or "h " — covers Excel typos
# like "3) DM", "61& Amazon", "16 h Esben", "14)€ Urmel".
AMOUNT_FALLBACK_RE = re.compile(r"^\s*(?P<amount>\d{1,6}(?:[.,]\d{1,2})?)\s*[)&h]\s*(€\s*)?\S")
# Date like "24.11." or "15.11."
INLINE_DATE_RE = re.compile(r"(?P<d>\d{1,2})\.(?P<m>\d{1,2})\.")


def parse_amount(s: str) -> float:
    return float(s.replace(".", "").replace(",", ".") if s.count(",") == 1 else s.replace(",", "."))


def parse_note_lines(note: str, month: int) -> list[dict]:
    """Return list of {amount_cents, description, day}."""
    out = []
    if not note:
        return out
    for raw_line in note.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        m = AMOUNT_RE.search(line)
        if not m:
            m = AMOUNT_FALLBACK_RE.match(line)
            if not m:
                continue
        amount = parse_amount(m.group("amount"))
        # description = line without the matched amount-euro token
        desc = (line[: m.start()] + line[m.end():]).strip(" -,;:\t")
        # prefer inline date if present
        day = DEFAULT_DAY
        dm = INLINE_DATE_RE.search(line)
        if dm:
            try:
                d = int(dm.group("d"))
                mm = int(dm.group("m"))
                if 1 <= d <= 31 and mm == month:
                    day = d
            except ValueError:
                pass
        # strip leading date token from description
        if dm:
            desc = (desc[: dm.start()] + desc[dm.end():]).strip(" -,;:\t")
        if not desc:
            desc = "(no description)"
        out.append({
            "amount_cents": int(round(amount * 100)),
            "description": desc,
            "day": day,
        })
    return out


# ---------- Plan types ----------

@dataclass
class PlannedTx:
    month: int
    day: int
    category: str | None  # category name or None
    is_income: bool
    description: str
    amount_cents: int  # signed: positive=income, negative=outcome

@dataclass
class PlannedRecurring:
    category: str | None
    description: str
    amount_cents: int  # signed
    day_of_month: int


@dataclass
class ImportPlan:
    transactions: list[PlannedTx] = field(default_factory=list)
    recurring: list[PlannedRecurring] = field(default_factory=list)
    new_categories: set[tuple[str, bool]] = field(default_factory=set)  # (name, isIncome)
    unparsed_lines: list[tuple[str, int, str]] = field(default_factory=list)  # (cell, month, line)


# ---------- Plan builder ----------

def _add_tx_from_note(plan: ImportPlan, note: str, month: int, category: str | None,
                      is_income: bool, label: str, prefix: str | None,
                      cell_ref: str):
    parsed = parse_note_lines(note, month)
    if parsed:
        for p in parsed:
            desc = p["description"]
            if prefix:
                desc = f"{prefix}: {desc}"
            plan.transactions.append(PlannedTx(
                month=month, day=p["day"], category=category, is_income=is_income,
                description=desc,
                amount_cents=p["amount_cents"] if is_income else -p["amount_cents"],
            ))
    # lines that had no amount at all (neither primary nor fallback) go to unparsed log
    for raw_line in (note or "").splitlines():
        line = raw_line.strip()
        if line and not AMOUNT_RE.search(line) and not AMOUNT_FALLBACK_RE.match(line):
            plan.unparsed_lines.append((cell_ref, month, line))


def build_plan(wb_path: str) -> ImportPlan:
    wb = load_workbook(wb_path, data_only=False)
    ws = wb["Tabellenblatt1"]
    plan = ImportPlan()

    # --- Income (Z. 2–5) ---
    # Cell values are authoritative; comments provide detail.
    # A reconciliation TX is added when parsed comments differ from cell value.
    for row, (label, cat, is_income, prefix) in INCOME_ROWS.items():
        plan.new_categories.add((cat, is_income))
        for month, col in MONTH_COLS.items():
            cell = ws[f"{col}{row}"]
            val = cell.value
            note = cell.comment.text if cell.comment else None
            cell_cents = int(round(float(val) * 100)) if isinstance(val, (int, float)) and val else 0
            if note and cell_cents:
                before = len(plan.transactions)
                _add_tx_from_note(plan, note, month, cat, is_income, label, prefix, f"{col}{row}")
                parsed_sum = sum(t.amount_cents for t in plan.transactions[before:])
                if parsed_sum != cell_cents:
                    diff = cell_cents - parsed_sum
                    plan.transactions.append(PlannedTx(
                        month=month, day=DEFAULT_DAY, category=cat, is_income=is_income,
                        description=f"{label} (Anpassung)", amount_cents=diff,
                    ))
            elif cell_cents:
                plan.transactions.append(PlannedTx(
                    month=month, day=DEFAULT_DAY, category=cat, is_income=is_income,
                    description=label, amount_cents=cell_cents,
                ))

    # --- Alltagsausgaben (Z. 8–26) ---
    # Same reconciliation: cell value is authoritative total per row/month.
    for row, (label, cat, is_income, prefix) in OUTCOME_ROWS.items():
        plan.new_categories.add((cat, is_income))
        for month, col in MONTH_COLS.items():
            cell = ws[f"{col}{row}"]
            val = cell.value
            note = cell.comment.text if cell.comment else None
            cell_cents = int(round(float(val) * 100)) if isinstance(val, (int, float)) and val else 0
            if note and cell_cents:
                before = len(plan.transactions)
                _add_tx_from_note(plan, note, month, cat, is_income, label, prefix, f"{col}{row}")
                parsed_sum = sum(t.amount_cents for t in plan.transactions[before:])
                expected = -cell_cents  # expenses are negative
                if parsed_sum != expected:
                    diff = expected - parsed_sum
                    plan.transactions.append(PlannedTx(
                        month=month, day=DEFAULT_DAY, category=cat, is_income=False,
                        description=f"{label} (Anpassung)", amount_cents=diff,
                    ))
            elif note and not cell_cents:
                # Comment exists but cell value is 0 → skip (cell is authoritative)
                pass
            elif cell_cents:
                plan.transactions.append(PlannedTx(
                    month=month, day=DEFAULT_DAY, category=cat, is_income=False,
                    description=label, amount_cents=-cell_cents,
                ))

    # --- Jährliche Posten (Z. 29–36) → Fixed costs ---
    plan.new_categories.add((FIXED_COST_CATEGORY, False))
    for row, label in YEARLY_ROWS.items():
        for month, col in MONTH_COLS.items():
            val = ws[f"{col}{row}"].value
            if isinstance(val, (int, float)) and val and val > 0:
                plan.transactions.append(PlannedTx(
                    month=month, day=DEFAULT_DAY, category=FIXED_COST_CATEGORY, is_income=False,
                    description=label, amount_cents=-int(round(float(val) * 100)),
                ))

    # --- Sparen (Z. 38–41) ---
    plan.new_categories.add(("Savings", False))
    for row, (label, prefix) in SAVINGS_ROWS.items():
        for month, col in MONTH_COLS.items():
            cell = ws[f"{col}{row}"]
            val = cell.value
            note = cell.comment.text if cell.comment else None
            if note:
                _add_tx_from_note(plan, note, month, "Savings", False, label, prefix, f"{col}{row}")
            elif isinstance(val, (int, float)) and val and val > 0:
                plan.transactions.append(PlannedTx(
                    month=month, day=DEFAULT_DAY, category="Savings", is_income=False,
                    description=f"{prefix}: {label}", amount_cents=-int(round(float(val) * 100)),
                ))

    # --- Monatliche Fixkosten (Z. 43–66) ---
    for row, (label, mode) in FIXED_COST_ROWS.items():
        values = {m: ws[f"{c}{row}"].value for m, c in MONTH_COLS.items()}
        numeric = {m: float(v) for m, v in values.items() if isinstance(v, (int, float))}

        if mode == "skip_zero" and all((v or 0) == 0 for v in numeric.values()):
            continue

        if mode == "single":
            # every non-zero month → single TX
            for m, v in numeric.items():
                if v and v > 0:
                    plan.transactions.append(PlannedTx(
                        month=m, day=DEFAULT_DAY, category=FIXED_COST_CATEGORY, is_income=False,
                        description=label, amount_cents=-int(round(v * 100)),
                    ))
            continue

        if mode == "recurring":
            # Use the latest month's value as recurring (for future projections).
            # Also create single TX for EVERY month so the Transaction table has
            # complete historical data (needed for correct carryover calculation).
            latest = numeric.get(max(numeric.keys())) if numeric else None
            if latest is None or latest <= 0:
                # no meaningful latest value → skip
                continue
            plan.recurring.append(PlannedRecurring(
                category=FIXED_COST_CATEGORY, description=label,
                amount_cents=-int(round(latest * 100)), day_of_month=1,
            ))
            for m, v in numeric.items():
                if v is None or v <= 0:
                    continue
                plan.transactions.append(PlannedTx(
                    month=m, day=DEFAULT_DAY, category=FIXED_COST_CATEGORY, is_income=False,
                    description=label, amount_cents=-int(round(v * 100)),
                ))

    return plan


# ---------- Reporting ----------

def print_plan_summary(plan: ImportPlan):
    print("=" * 70)
    print("IMPORT PLAN — Haushalt 2026.xlsx → acc_demo")
    print("=" * 70)

    print(f"\nNew categories to create (skip if already present):")
    for name, is_inc in sorted(plan.new_categories):
        print(f"  - {name}  (income={is_inc})")

    print(f"\nRecurring transactions: {len(plan.recurring)}")
    for r in plan.recurring:
        sign = "+" if r.amount_cents >= 0 else "-"
        print(f"  - {sign}{abs(r.amount_cents)/100:>8.2f} € | {r.category:<12} | {r.description}")

    print(f"\nSingle transactions: {len(plan.transactions)}")
    by_month_cat = defaultdict(lambda: defaultdict(lambda: {"count": 0, "in": 0, "out": 0}))
    for t in plan.transactions:
        entry = by_month_cat[t.month][t.category or "—"]
        entry["count"] += 1
        if t.amount_cents >= 0:
            entry["in"] += t.amount_cents
        else:
            entry["out"] += -t.amount_cents

    for m in sorted(by_month_cat):
        total_in = sum(c["in"] for c in by_month_cat[m].values())
        total_out = sum(c["out"] for c in by_month_cat[m].values())
        print(f"\n  -- Month {m:02d}/{YEAR} -- total {sum(c['count'] for c in by_month_cat[m].values())} TX"
              f"   income={total_in/100:.2f}€  outcome={total_out/100:.2f}€")
        for cat, v in sorted(by_month_cat[m].items()):
            parts = []
            if v["in"]:
                parts.append(f"+{v['in']/100:.2f}€")
            if v["out"]:
                parts.append(f"-{v['out']/100:.2f}€")
            print(f"     {cat:<18}  n={v['count']:>3}  {', '.join(parts)}")

    if plan.unparsed_lines:
        print(f"\nUnparsed note lines (skipped, review if needed): {len(plan.unparsed_lines)}")
        for cell, month, line in plan.unparsed_lines[:30]:
            print(f"  {cell} (m={month}): {line!r}")
        if len(plan.unparsed_lines) > 30:
            print(f"  ... and {len(plan.unparsed_lines) - 30} more")


# ---------- DB apply ----------

def apply_plan(plan: ImportPlan):
    conn = psycopg2.connect(DB_URL)
    conn.autocommit = False
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            # Resolve user id
            cur.execute('SELECT id FROM "User" WHERE email = %s', (USER_EMAIL,))
            user = cur.fetchone()
            if not user:
                raise RuntimeError(f"user {USER_EMAIL} not found")
            user_id = user["id"]

            # Wipe demo data for acc_demo
            cur.execute('DELETE FROM "RecurringTransactionSkip" WHERE "recurringId" IN '
                        '(SELECT id FROM "RecurringTransaction" WHERE "accountId" = %s)', (ACCOUNT_ID,))
            cur.execute('DELETE FROM "RecurringTransaction" WHERE "accountId" = %s', (ACCOUNT_ID,))
            cur.execute('DELETE FROM "Transaction" WHERE "accountId" = %s', (ACCOUNT_ID,))
            cur.execute('DELETE FROM "Budget" WHERE "accountId" = %s', (ACCOUNT_ID,))

            # Ensure categories exist
            cur.execute('SELECT id, name, "isIncome" FROM "Category" WHERE "userId" = %s', (user_id,))
            cats = {(r["name"], r["isIncome"]): r["id"] for r in cur.fetchall()}
            for name, is_inc in plan.new_categories:
                if (name, is_inc) in cats:
                    continue
                # sometimes same name exists with different isIncome flag; our unique is (userId, name)
                # treat name as authoritative: if exists with other flag, update flag
                existing_same_name = [v for (n, _), v in cats.items() if n == name]
                if existing_same_name:
                    # update isIncome to match planned value
                    cur.execute('UPDATE "Category" SET "isIncome" = %s WHERE id = %s',
                                (is_inc, existing_same_name[0]))
                    cats[(name, is_inc)] = existing_same_name[0]
                    continue
                new_id = _cuid_like()
                cur.execute('INSERT INTO "Category" (id, name, "userId", "isIncome") VALUES (%s,%s,%s,%s)',
                            (new_id, name, user_id, is_inc))
                cats[(name, is_inc)] = new_id

            # Insert transactions
            for t in plan.transactions:
                cat_id = None
                if t.category:
                    cat_id = cats.get((t.category, t.is_income))
                    if cat_id is None:
                        # fallback: match any isIncome flag with same name
                        cat_id = next((v for (n, _), v in cats.items() if n == t.category), None)
                occurred = datetime(YEAR, t.month, min(t.day, _last_day_of(t.month)))
                new_id = _cuid_like()
                cur.execute(
                    'INSERT INTO "Transaction" (id,"accountId","categoryId","amountCents","description","occurredAt") '
                    'VALUES (%s,%s,%s,%s,%s,%s)',
                    (new_id, ACCOUNT_ID, cat_id, t.amount_cents, t.description, occurred),
                )

            # Insert recurring transactions
            for r in plan.recurring:
                cat_id = None
                if r.category:
                    cat_id = next((v for (n, _), v in cats.items() if n == r.category), None)
                # next occurrence = 1st of next month at day_of_month
                now = datetime.now()
                ny, nm = (now.year, now.month + 1) if now.month < 12 else (now.year + 1, 1)
                dom = min(r.day_of_month, _last_day_of_ym(ny, nm))
                next_occ = datetime(ny, nm, dom)
                new_id = _cuid_like()
                cur.execute(
                    'INSERT INTO "RecurringTransaction" '
                    '(id,"accountId","categoryId","amountCents","description",frequency,"intervalMonths","dayOfMonth","nextOccurrence") '
                    'VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)',
                    (new_id, ACCOUNT_ID, cat_id, r.amount_cents, r.description,
                     "MONTHLY", 1, r.day_of_month, next_occ),
                )
        conn.commit()
        print("✅ applied.")
    except Exception:
        conn.rollback()
        print("❌ rolled back.")
        raise
    finally:
        conn.close()


def _last_day_of(month: int) -> int:
    return _last_day_of_ym(YEAR, month)


def _last_day_of_ym(y: int, m: int) -> int:
    from calendar import monthrange
    return monthrange(y, m)[1]


def _cuid_like() -> str:
    # Prisma-compatible id shape is not strictly required by Postgres, but keep it string-safe.
    import secrets
    import string
    alphabet = string.ascii_lowercase + string.digits
    return "xlsx_" + "".join(secrets.choice(alphabet) for _ in range(20))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--apply", action="store_true")
    args = ap.parse_args()
    if not args.dry_run and not args.apply:
        print("Pass --dry-run or --apply", file=sys.stderr)
        sys.exit(2)
    plan = build_plan(XLSX_PATH)
    print_plan_summary(plan)
    if args.apply:
        print("\n--- APPLYING ---")
        apply_plan(plan)


if __name__ == "__main__":
    main()
